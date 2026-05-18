"""Load camera intrinsics and ArUco marker 3D coordinates from config files."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import numpy as np
import numpy.typing as npt
import yaml


def load_intrinsics_yaml(path: Path) -> dict[str, Any]:
    """Load camera intrinsics and timing parameters from a YAML file."""
    with path.open(encoding="utf-8") as handle:
        data = yaml.safe_load(handle)
    if not isinstance(data, dict):
        raise ValueError(f"Invalid intrinsics config: {path}")
    return data


def intrinsics_matrix_from_fov(
    width: int,
    height: int,
    fov_x_deg: float,
    fov_y_deg: float,
) -> npt.NDArray[np.float64]:
    """Build a 3x3 pinhole intrinsic matrix K from horizontal/vertical field of view."""
    fov_x = np.deg2rad(fov_x_deg)
    fov_y = np.deg2rad(fov_y_deg)
    fx = (width / 2.0) / np.tan(fov_x / 2.0)
    fy = (height / 2.0) / np.tan(fov_y / 2.0)
    cx = width / 2.0
    cy = height / 2.0
    return np.array(
        [[fx, 0.0, cx], [0.0, fy, cy], [0.0, 0.0, 1.0]],
        dtype=np.float64,
    )


def scale_intrinsics(
    camera_matrix: npt.NDArray[np.float64],
    config_width: int,
    config_height: int,
    video_width: int,
    video_height: int,
) -> npt.NDArray[np.float64]:
    """Scale focal length and principal point when video resolution differs from config."""
    sx = video_width / config_width
    sy = video_height / config_height
    scaled = camera_matrix.copy()
    scaled[0, 0] *= sx
    scaled[1, 1] *= sy
    scaled[0, 2] *= sx
    scaled[1, 2] *= sy
    return scaled


def build_camera_matrix(
    intrinsics: dict[str, Any],
    video_width: int,
    video_height: int,
) -> npt.NDArray[np.float64]:
    """Build K for the given video frame size."""
    config_w = int(intrinsics["image_width"])
    config_h = int(intrinsics["image_height"])
    k = intrinsics_matrix_from_fov(
        config_w,
        config_h,
        float(intrinsics["fov_x_deg"]),
        float(intrinsics["fov_y_deg"]),
    )
    if intrinsics.get("scale_intrinsics_to_video", True):
        if (config_w, config_h) != (video_width, video_height):
            k = scale_intrinsics(k, config_w, config_h, video_width, video_height)
    return k


def distortion_coefficients(intrinsics: dict[str, Any]) -> npt.NDArray[np.float64]:
    """Return OpenCV distortion vector from config."""
    coeffs = intrinsics.get("distortion", [0.0, 0.0, 0.0, 0.0, 0.0])
    return np.asarray(coeffs, dtype=np.float64).reshape(-1)


def load_markers_3d(path: Path) -> dict[int, npt.NDArray[np.float64]]:
    """
    Load wall ArUco marker upper-right corner positions (mm) keyed by marker ID.

    Supports:
    - Excel: ``ArUco_markers_3D.xlsx`` (columns Marker_ID, X, Y, Z)
    - CSV with header: marker_id, x_mm, y_mm, z_mm (flexible names)
    """
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return load_markers_3d_xlsx(path)
    if suffix == ".csv":
        return load_markers_3d_csv(path)
    raise ValueError(f"Unsupported markers file type: {path}")


def load_markers_3d_xlsx(path: Path) -> dict[int, npt.NDArray[np.float64]]:
    """Load marker positions from DPJAIT ``ArUco_markers_3D.xlsx``."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read .xlsx marker files: pip install openpyxl",
        ) from exc

    positions: dict[int, npt.NDArray[np.float64]] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"Empty markers workbook: {path}")

        header_map = {
            str(cell).strip().lower(): idx
            for idx, cell in enumerate(header)
            if cell is not None
        }
        id_idx = header_map.get("marker_id", header_map.get("id", 0))
        x_idx = header_map.get("x", header_map.get("x_mm", 1))
        y_idx = header_map.get("y", header_map.get("y_mm", 2))
        z_idx = header_map.get("z", header_map.get("z_mm", 3))

        for row in rows:
            if row is None or row[id_idx] is None:
                continue
            marker_id = int(float(row[id_idx]))
            positions[marker_id] = np.array(
                [float(row[x_idx]), float(row[y_idx]), float(row[z_idx])],
                dtype=np.float64,
            )
    finally:
        workbook.close()

    if not positions:
        raise ValueError(f"No marker rows loaded from {path}")
    return positions


def load_markers_3d_csv(path: Path) -> dict[int, npt.NDArray[np.float64]]:
    """
    Load marker upper-right corner positions (mm) keyed by marker ID.

    Expected columns: marker_id, x_mm, y_mm, z_mm (header names are flexible).

    Note: DPJAIT ``*_U.csv`` files are mocap trajectories, not wall markers.
    Use ``--mocap`` for those and ``ArUco_markers_3D.xlsx`` for wall marker coords.
    """
    positions: dict[int, npt.NDArray[np.float64]] = {}
    with path.open(encoding="utf-8", newline="") as handle:
        lines = [line for line in handle if not line.lstrip().startswith("#")]
        if not lines:
            raise ValueError(f"Empty markers file: {path}")

        reader = csv.reader(lines)
        first = next(reader)
        if _looks_like_dpjait_mocap_row(first):
            raise ValueError(
                f"{path} looks like a DPJAIT mocap trajectory (*_U.csv), not wall markers. "
                "Use --mocap for this file and --markers data/raw/ArUco_markers_3D.xlsx",
            )

        header_lower = [cell.strip().lower() for cell in first]
        has_header = any(
            name in header_lower
            for name in ("marker_id", "id", "x", "x_mm", "markerid")
        )

        def parse_row(cells: list[str]) -> None:
            if len(cells) < 4:
                return
            marker_id = int(float(cells[0]))
            positions[marker_id] = np.array(
                [float(cells[1]), float(cells[2]), float(cells[3])],
                dtype=np.float64,
            )

        if has_header:
            dict_reader = csv.DictReader(lines)
            if dict_reader.fieldnames is None:
                raise ValueError(f"No header row in markers file: {path}")

            fields = {name.strip().lower(): name for name in dict_reader.fieldnames}

            def pick(*candidates: str) -> str:
                for key in candidates:
                    if key in fields:
                        return fields[key]
                raise ValueError(
                    f"Missing column in {path}. Need marker id and x,y,z; "
                    f"got {dict_reader.fieldnames}",
                )

            id_col = pick("marker_id", "id", "markerid", "aruco_id")
            x_col = pick("x_mm", "x", "pos_x")
            y_col = pick("y_mm", "y", "pos_y")
            z_col = pick("z_mm", "z", "pos_z")

            for row in dict_reader:
                marker_id = int(float(row[id_col]))
                positions[marker_id] = np.array(
                    [float(row[x_col]), float(row[y_col]), float(row[z_col])],
                    dtype=np.float64,
                )
        else:
            parse_row(first)
            for row in reader:
                parse_row(row)

    if not positions:
        raise ValueError(f"No marker rows loaded from {path}")
    return positions


def _looks_like_dpjait_mocap_row(cells: list[str]) -> bool:
    """Heuristic: mocap rows start with index, 0, then many numeric columns."""
    if len(cells) < 14:
        return False
    try:
        int(float(cells[0]))
        int(float(cells[1]))
        floats = [float(c) for c in cells[2:]]
    except ValueError:
        return False
    return len(floats) >= 12 and len(floats) % 3 == 0
